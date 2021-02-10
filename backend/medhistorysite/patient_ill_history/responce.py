import requests
import json
import openpyxl

def get():
    with requests.get('http://127.0.0.1:8000/api/PatientAll/?format=json') as response:
        return json.loads(response.text)
# def get():
#     with open('data.json') as file:
#         return json.load(file)


book = openpyxl.Workbook()
sheet = book.active

sheet.column_dimensions['A'].width = 15  # NAME
sheet.column_dimensions['B'].width = 15  # SURNAME
sheet.column_dimensions['C'].width = 35  # TREATMENT_SESSION_startSession
sheet.column_dimensions['D'].width = 70  # TREATMENT_SESSION_mainIll
sheet.column_dimensions['E'].width = 30  # TREATMENT_SESSION_doctor
sheet.column_dimensions['F'].width = 70  # TREATMENT_SESSION_comorbidity
sheet.column_dimensions['G'].width = 70  # STAGE_OF_TREATMENT_stageName
sheet.column_dimensions['H'].width = 70  # STAGE_OF_TREATMENT_startStage
sheet.column_dimensions['I'].width = 90  # STAGE_OF_TREATMENT_surgery
sheet.column_dimensions['J'].width = 90  # STAGE_OF_TREATMENT_pharmacotherapy
sheet.column_dimensions['K'].width = 70  # STAGE_OF_TREATMENT_physiotherapy
sheet.column_dimensions['L'].width = 70  # STAGE_OF_TREATMENT_electro_ultrasound_therapy
sheet.column_dimensions['M'].width = 30  # STAGE_OF_TREATMENT_stateOn
sheet.column_dimensions['N'].width = 200  # STAGE_OF_TREATMENT_state_laboratory_tests
sheet.column_dimensions['O'].width = 200  # STAGE_OF_TREATMENT_state_measurement
sheet.column_dimensions['P'].width = 60  # STAGE_OF_TREATMENT_endStage
sheet.column_dimensions['Q'].width = 60  # STAGE_OF_TREATMENT_endSession
sheet.column_dimensions['R'].width = 100
sheet.column_dimensions['S'].width = 100
sheet.column_dimensions['T'].width = 100

row = 2
for patient in get():
    sheet.cell(1, 1, 'NAME')
    sheet[row][0].value = patient['name']
    sheet.cell(1, 2, 'SURNAME')
    sheet[row][1].value = patient['surname']
    i = 0
    indCel = 0
    for treatment_session in patient['treatment_session']:
        no_of_session = i+1
        sheet.cell(1, indCel+3, ('TREATMENT_SESSION_' + str(no_of_session) + '_STARTSESSION_'))
        sheet[row][indCel+2].value = treatment_session['startSession']
        sheet.cell(1, indCel + 4, ('TREATMENT_SESSION_' + str(no_of_session) + '_MAINILL' ))
        sheet[row][indCel+3].value = treatment_session['mainIll']
        sheet.cell(1, indCel + 5, ('TREATMENT_SESSION_' + str(no_of_session) + '_DOCTOR'))
        sheet[row][indCel+4].value = treatment_session['doctor']

        comorbidity_str = ''
        for comorbidity in treatment_session['comorbidity']:
            comorbidity_str += comorbidity['nameIll'] + ' '
        sheet.cell(1, indCel + 6, ('TREATMENT_SESSION_' + str(no_of_session) + '_COMORBIDITY'))
        sheet[row][indCel+5].value = comorbidity_str

        a = 0
        for stage_of_treatment in treatment_session['stage_of_treatment']:
            sheet.cell(1, indCel + a + 7, ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_STAGENAME_'))
            sheet[row][indCel+a+6].value = stage_of_treatment['stageName']
            sheet.cell(1, indCel + a + 8, ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_STARTSTAGE_'))
            sheet[row][indCel+a+7].value = stage_of_treatment['startStage']

            surgery_str = ''
            for surgery in stage_of_treatment['surgery']:
                surgery_str += surgery['DateIntervention'] + ' ' + surgery['nameInterven'] + ' '
                for sur_med_staff in surgery['sur_med_staff']:
                    surgery_str += sur_med_staff['medStaff']
            sheet.cell(1, indCel + a + 9, ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_ _SURGERY'))
            sheet[row][indCel+a+8].value = surgery_str

            pharmacotherapy_str = ''
            for pharmacotherapy in stage_of_treatment['pharmacotherapy']:
                pharmacotherapy_str += pharmacotherapy['namePill'] + ' ' + str(pharmacotherapy['dosePill']) + ' ' + \
                    str(pharmacotherapy['unitPill']) + ' ' + str(pharmacotherapy['datePill']) + ' '
            sheet.cell(1, indCel + a + 10, ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_ _pharmacotherapy'))
            sheet[row][indCel+a+9].value = pharmacotherapy_str

            physiotherapy_str = ''
            for physiotherapy in stage_of_treatment['physiotherapy']:
                physiotherapy_str += physiotherapy['namePhysiotherapy'] + ' ' + str(physiotherapy['valuePhysiotherapy']) + ' ' +\
                str(physiotherapy['unitPhysiotherapy']) + ' ' + str(physiotherapy['datePhysiotherapy']) + ' '
                for physiotherapy_med_staff in physiotherapy['physiotherapy_med_staff']:
                    physiotherapy_str += physiotherapy_med_staff['medStaff']
            sheet.cell(1, indCel + a + 11,
                       ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_ _physiotherapy'))
            sheet[row][indCel+a+10].value = physiotherapy_str

            for electro_ultrasound_therapy in stage_of_treatment['electro_ultrasound_therapy']:
                continue
            sheet.cell(1, indCel + a + 12,
                           ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_' + '_electro_ultrasound_therapy'))

            for state in stage_of_treatment['state']:
                sheet.cell(1, indCel + a + 13,
                           ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_ _state'))
                sheet[row][indCel+a+12].value = state['state_on']

                laboratory_test_str = ''
                for laboratory_test in state['laboratory_test']:
                    laboratory_test_str += str(laboratory_test['nameTest']) + ' ' + str(laboratory_test['valueTest']) + ' ' + \
                        str(laboratory_test['unitTest']) + ' ' + str(laboratory_test['dateTest']) + ' ' + \
                        str(laboratory_test['laboratoryName']) + ' '
                    for lab_staff in laboratory_test['lab_staff']:  # test
                        laboratory_test_str += lab_staff['id_lab_test'] + ' ' + lab_staff['id_med_staff'] + ' '

                sheet.cell(1, indCel + a + 14,
                           ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_ _laboratory_test'))
                sheet[row][indCel+a+13].value = laboratory_test_str

                measurement_str = ''
                for measurement in state['measurement']:
                    measurement_str += str(measurement['nameMeasurement']) + ' ' + str(measurement['valueMeasurement']) + ' ' + \
                        str(measurement['unitMeasurement']) + ' ' + str(measurement['dateMeasurement']) + ' ' + \
                        str(measurement['medStaff']) + ' '
                sheet.cell(1, indCel + a + 15,
                           ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_' + '_measurement'))
                sheet[row][indCel+a+14].value = measurement_str

                for image in state['image']:
                    pass

            sheet.cell(1, indCel + a + 16,
                       ('TREATMENT_SESSION_' + str(no_of_session) + '_STAGE_OF_TREATMENT_' + '_endStage'))
            sheet[row][indCel+a+15].value = stage_of_treatment['endStage']
            a = a + 10

        sheet.cell(1, indCel + a + 7, ('TREATMENT_SESSION_' + str(no_of_session) + '_endSession'))
        sheet[row][indCel + a +6].value = treatment_session['endSession']
        i+=1
        indCel = indCel + 5 + a
row += 1


def as_text(value):
    if value is None:
        return ""
    return str(value)

for column_cells in sheet.columns:
    length = max(len(as_text(cell.value)) for cell in column_cells)
    sheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length+30


book.save('test.xlsx')
book.close()

print("Finish")
